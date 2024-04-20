import csv
import datetime
import json
import logging
import os
import sqlite3
import uuid
from hashlib import sha256
from pathlib import Path
from sqlite3 import Error
from threading import Thread

import nltk
from nltk.tokenize import sent_tokenize
import pdfplumber
from docx import Document
import openpyxl

logging.basicConfig(level=logging.INFO)


class DocumentLoader:
    FILE_TYPES = {
        '.pdf': 'pdf',
        '.json': 'json',
        '.txt': 'txt',
        '.docx': 'docx',
        '.xlsx': 'xlsx',
        '.csv': 'csv'
    }

    def __init__(self, storage_folder="../storage", db_file="../knowledge.db"):
        self.storage_folder = Path(storage_folder)
        self.db_file = db_file
        self.init_db()

    def connect_db(self):
        try:
            return sqlite3.connect(self.db_file)
        except Error as e:
            logging.error(f"Failed to connect to the database: {e}")
            return None

    def init_db(self):
        conn = self.connect_db()
        if conn:
            with conn:
                self.create_tables(conn.cursor())

    @staticmethod
    def generate_hash(filepath):
        hash_alg = sha256()
        with open(filepath, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_alg.update(chunk)
        return hash_alg.hexdigest()

    @staticmethod
    def extract_text_from_pdf(filepath):
        text = []
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text.append(page.extract_text())
        return ''.join(text)

    @staticmethod
    def extract_text_from_txt(filepath):
        with open(filepath, 'r') as f:
            return f.read()

    @staticmethod
    def extract_text_from_docx(filepath):
        doc = Document(filepath)
        return '\n'.join([paragraph.text for paragraph in doc.paragraphs])

    @staticmethod
    def extract_text_from_xlsx(filepath):
        workbook = openpyxl.load_workbook(filepath)
        text = []
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows():
                row_text = [str(cell.value) for cell in row if cell.value is not None]
                text.append(' '.join(row_text))
        return '\n'.join(text)

    @staticmethod
    def extract_text_from_csv(filepath):
        def read_csv(file_path, encoding):
            text = []
            with open(file_path, newline='', encoding=encoding) as csvfile:
                reader = csv.reader(csvfile)
                for row in reader:
                    row_text = ', '.join(cell for cell in row if cell)
                    text.append(row_text)
            return '\n'.join(text)

        try:
            return read_csv(filepath, 'utf-8')
        except UnicodeDecodeError:
            logging.info(f"UTF-8 decoding failed for {filepath}, trying Windows-1252")
            return read_csv(filepath, 'windows-1252')

    @staticmethod
    def create_tables(cursor):
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS documents (
                id TEXT PRIMARY KEY,
                filename TEXT,
                filetype TEXT,
                hash TEXT UNIQUE,
                created_at TEXT
            );
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS content (
                id TEXT PRIMARY KEY,
                document_id TEXT,
                content TEXT,
                sentence_number INTEGER,
                FOREIGN KEY (document_id) REFERENCES documents (id)
            );
        """)

    def load_documents(self):
        threads = []
        for filepath in self.storage_folder.iterdir():
            if filepath.is_file():
                thread = Thread(target=self.process_document, args=(filepath,))
                threads.append(thread)
                thread.start()

        for thread in threads:
            thread.join()

    def process_document(self, filepath):
        try:
            filetype = self.determine_filetype(filepath.name)
            if not filetype:
                logging.warning(f"No supported filetype found for {filepath}")
                return

            file_hash = self.generate_hash(filepath)
            if self.check_duplicate(file_hash):
                logging.info(f"Duplicate file skipped: {filepath}")
                return

            text = self.extract_text(filepath, filetype)
            if text is not None:
                self.store_document(filepath.name, filetype, file_hash, text)

            logging.info(f"File added: {filepath}")
        except Exception as e:
            logging.error(f"Error processing file {filepath}: {e}")

    def determine_filetype(self, filename):
        extension = Path(filename).suffix.lower()
        return self.FILE_TYPES.get(extension)

    def check_duplicate(self, file_hash):
        conn = self.connect_db()
        if conn:
            with conn:
                cursor = conn.cursor()
                cursor.execute("SELECT 1 FROM documents WHERE hash = ?", (file_hash,))
                return cursor.fetchone() is not None

    def extract_text(self, filepath, filetype):
        if filetype == 'pdf':
            return self.extract_text_from_pdf(filepath)
        elif filetype == 'json':
            return self.extract_text_from_json(filepath)
        elif filetype == 'txt':
            return self.extract_text_from_txt(filepath)
        elif filetype == 'docx':
            return self.extract_text_from_docx(filepath)
        elif filetype == 'xlsx':
            return self.extract_text_from_xlsx(filepath)
        elif filetype == 'csv':
            return self.extract_text_from_csv(filepath)
        else:
            logging.error(f"Unsupported file type: {filetype}")
            return None

    def store_document(self, filename, filetype, file_hash, text):
        conn = self.connect_db()
        if conn:
            with conn:
                cursor = conn.cursor()
                document_id = str(uuid.uuid4())
                cursor.execute(
                    "INSERT INTO documents (id, filename, filetype, hash, created_at) VALUES (?, ?, ?, ?, ?)",
                    (document_id, filename, filetype, file_hash, datetime.datetime.now().isoformat()))
                for i, sentence in enumerate(sent_tokenize(text)):
                    content_id = str(uuid.uuid4())
                    cursor.execute(
                        "INSERT INTO content (id, document_id, content, sentence_number) VALUES (?, ?, ?, ?)",
                        (content_id, document_id, sentence, i + 1))

    def extract_text_from_json(self, filepath):
        with open(filepath, 'r') as f:
            data = json.load(f)
        return '\n'.join(f"{k}: {v}" for k, v in self.traverse_json(data))

    def traverse_json(self, obj, prefix=''):
        if isinstance(obj, dict):
            for key, value in obj.items():
                yield from self.traverse_json(value, prefix + key + '.')
        elif isinstance(obj, list):
            for i, item in enumerate(obj):
                yield from self.traverse_json(item, prefix + f'[{i}].')
        else:
            yield prefix[:-1], str(obj)  # remove the trailing dot


if __name__ == "__main__":
    loader = DocumentLoader()
    loader.load_documents()
